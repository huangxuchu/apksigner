"""
为apk包签名。
需求：
1、批量签名apk
2、多个签名文件可选
逻辑：
1、网页上传apk文件到本地的temp文件夹下，并生成时间戳的文件夹。没有temp文件夹时需要自动生成。
2、签名程序获取本地temp目录下的时间错文件夹的目录下的所有apk
"""
import json
import os
import sys

from openpyxl.utils import get_column_letter

import CmdUtils
import Config
import Constants
import DateUtils
import ExcelFileUtils
import FileUtils
import LogUtils
from ApkInfo import ApkInfo

TAG = 'SignApk'

# 签名命令 参数：1=签名.jks或.keystore文件路径, 2=签名文件密码, 3=别名, 4=别名密码, 5=apk文件路径
APKSIGNER_ORDER_SIGN = """ sign --ks %s --ks-pass pass:%s --ks-key-alias %s --key-pass pass:%s %s"""
# 获取apk包签名md5 参数：1=apk文件路径
KEYTOOL_ORDER_APK_SIGN_INFO = """keytool -printcert -jarfile %s"""
# 获取keystore的md5 参数：1=keystore文件路径, 2=keystore的密码
KEYTOOL_ORDER_KEYSTORE_INFO = """keytool -list -v -keystore %s -storepass %s"""

# 签名信息 key=包名
_keystore_json = {}

# 正在执行的任务
signing_task = []


# selected_keystore: apk的签名文件
# task_id: 任务id
# input_path: 批量签名存放任务的目录
# output: 签名后的输出目录
def sign_batch(selected_keystore, task_id, input_path, output_path):
    if task_id is None or len(task_id) == 0:
        task_id = DateUtils.date_time()
    print('签名 task=' + task_id)
    print('Using Apksigner ' + (CmdUtils.popen(Config.APKSIGNER + ' --version')[0]).strip())
    if task_id in signing_task:
        return
    signing_task.append(task_id)
    taskPath = os.path.join(input_path)
    apkList = os.listdir(taskPath)
    apkList.sort()
    for apk in apkList:
        if not apk.endswith(".apk"):
            continue
        apkPath = os.path.join(taskPath, apk)
        code = _sign(selected_keystore, apkPath)
        if code == 0:
            apk_sign = apk.replace(".apk", "_sign.apk")
            os.renames(apkPath, os.path.join(output_path, task_id, apk_sign))
            os.remove(apkPath + r'.idsig')
    signing_task.remove(task_id)
    baseName = os.path.basename(input_path)
    if baseName != 'output' and input_path != output_path:
        FileUtils.rmdir(input_path)


def sign(selected_keystore, task_id, apk_path, output_path):
    """
    签名方法，签名后apk名称变为 xxx.apk->xxx_sign.apk
    :param selected_keystore: apk的签名文件
    :param task_id: 任务id
    :param apk_path: apk的目录
    :param output_path: 签名后的输出目录
    :return:
    """
    if task_id is None or len(task_id) == 0:
        task_id = DateUtils.date_time()
    print('开始签名 task=' + task_id)
    print('Using Apksigner ' + (CmdUtils.popen(Config.APKSIGNER + ' --version')[0]).strip())
    if not apk_path.endswith(".apk"):
        return
    signing_patch = apk_path + ".signing"
    FileUtils.copy(apk_path, signing_patch)
    code = _sign(selected_keystore, signing_patch)
    if code == 0:
        apk_signed = os.path.basename(signing_patch).replace(".apk.signing", "_sign.apk")
        os.renames(signing_patch, os.path.join(output_path, task_id, apk_signed))
        os.remove(signing_patch + r'.idsig')
    else:
        LogUtils.e(TAG, f'签名失败 code={code}')


def _sign(selected_keystore, apk_path):
    _init_keystore()
    for v in _keystore_json.values():
        print("遍历keystore" + str(v))
    keystore = None
    if selected_keystore is None or len(selected_keystore) == 0:
        # TODO 获取apk里面的包名，
        package_name = get_apk_info(apk_path)
        keystore = _keystore_json[package_name]
    else:
        # TODO 循环检查列表里是否有签名，没有就报错，有就执行
        for v in _keystore_json.values():
            if v[Constants.KEYSTORE] == selected_keystore:
                keystore = v
                break
    if keystore is None or len(keystore) == 0:
        LogUtils.e(TAG,
                   '未获取到签名数据，请检查"keystore"文件夹下相应的签名文件是否存在，以及确认签名配置"keystore.json"是否正确')
        return
    print("签名数据: " + keystore[Constants.APP_NAME] + " Apk路径: " + apk_path)
    task = Config.APKSIGNER + APKSIGNER_ORDER_SIGN % (
        os.path.join(Config.NOVEL_KEYSTORE_PATH, keystore[Constants.KEYSTORE]),
        keystore[Constants.PASSWORD],
        keystore[Constants.ALIAS],
        keystore[Constants.ALIAS_PASSWORD],
        apk_path)
    code = os.system(task)
    return code


def get_apk_info(apk_path):
    # TODO 返回包信息
    return ApkInfo("", "", "")


def _init_keystore():
    global _keystore_json
    if len(_keystore_json) == 0:
        sheet = ExcelFileUtils.get_sheet(Config.NOVEL_KEYSTORE_EXECL_PATH, ExcelFileUtils.DEFAULT_SHEET_ONE)
        highestRow = sheet.max_row
        highestColumn = sheet.max_column
        c = get_column_letter(highestColumn)
        c = c + str(highestRow)
        print("加载execl数据")
        obj = {}
        for rowOfCellObjects in sheet["A2":"G50"]:
            element = {}
            packageName = ""
            for cellObj in rowOfCellObjects:
                coordinate = cellObj.coordinate
                value = ""
                if cellObj.value is not None:
                    value = cellObj.value
                if coordinate.startswith("A"):
                    element[Constants.APP_NAME] = value
                elif coordinate.startswith("B"):
                    if value is None or len(value) <= 0:
                        break
                    element[Constants.PACKAGE_NAME] = value
                    packageName = value
                elif coordinate.startswith("C"):
                    element[Constants.KEYSTORE] = value
                elif coordinate.startswith("D"):
                    element[Constants.PASSWORD] = value
                elif coordinate.startswith("E"):
                    element[Constants.ALIAS] = value
                elif coordinate.startswith("F"):
                    element[Constants.ALIAS_PASSWORD] = value
                elif coordinate.startswith("G"):
                    element[Constants.REMARK] = value
                print(coordinate + "->" + value)
            if len(element) > 0 and Constants.PACKAGE_NAME in element:
                obj[packageName] = element
        _keystore_json = obj


def _parse_dc(info):
    """
    解析指纹证书的字符串
    :param info:
    :return:
    """
    key = "证书指纹:"
    start = False
    DC = []
    for line in info:
        line = line.strip()
        if start:
            if line.startswith("MD5:"):
                DC.append(line)
            elif line.startswith("SHA1:"):
                DC.append(line)
            elif line.startswith("SHA256:"):
                DC.append(line)
                break
        if not start:
            start = line.startswith(key)
    return DC


if __name__ == "__main__":
    print("启动SignApk")
    external_param_name = "/Users/hongxiang/Downloads/app-fantuankanshupro-release_120_jiagu.apk"
    for arg in sys.argv:
        print(arg)
        if arg.startswith('-a'):
            external_param_name = arg.replace('-a', '') + "-"
    sign("fantuantanshujbk.jks", None, external_param_name, "/Users/hongxiang/Downloads")
